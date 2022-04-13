# -*- coding: utf-8 -*-
# @Time    : 2022/3/8 15:59
# @Author  : kl
# @File    : excel_op.py
# 1、导入load_workbook
from openpyxl import load_workbook
# 2、加载一个本地的excel文件,得到一个工作薄对象
wb = load_workbook(r"E:\python projects\lemon\0308\testdatas\cases.xlsx")
# 3、通过工作薄,得到一个要操作的表单对象
sh = wb["注册"]
#4、获取单元格的数据-读
#行和列的下标都是从1开始的。行和列都是数字,# sh.cell(行号,列号).value
# cell_value = sh.cell(2, 2).value
# print(cell_value)

#修改/添加单元格的数据-写# sh.cell(行号,列号,重新给单元格赋值value)
# sh.cell(2,2, "py43")
# print(sh.cell(2, 2).value)
#修改之后,保存当前工作薄wb。这样子修改才会保存到文件当中去。# wb.save (文件名)
# 另存在(与打开的excel名字不一样或者在不同的路径)
#保存到原文件(打开excel时的原文件路径)
#注意:保存时,文件一定要先关闭掉,不要被其它程序所占用。
# wb.save ("case2.xlsx")

#读取一个表单当中所有的数据#1、如果单元格当中只有数字,就会转成整数类型
# 2、如果单元格当中只有TRUE、FALŞE,就会转成布尔值类型
# 3、如果单元格当中只有时间格式,比如2021/08/07,就会转成datetime类型
# for row_values in list(sh.values):
# 	print(row_values)

sh_all_datas = list(sh.values)
# key = sh_all_datas[0]
# print(key)
# zip函数使用:将第一行的key和第二行的value拼接成字典
# one_case = dict(zip(key, sh_all_datas [1]))
# print(one_case)
for values in sh_all_datas[1:]:
    print(values)

